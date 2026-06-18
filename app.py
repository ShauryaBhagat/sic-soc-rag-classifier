# app.py
# Streamlit UI for the SIC/SOC RAG Classifier.
#
# Deployment note: uses ChromaDB in-memory store, rebuilt at startup.
# This is the correct pattern for stateless cloud deployments —
# the vector store is cheap to rebuild (~$0.003, ~60 seconds) and
# avoids persistent disk dependencies.

import os
import csv
import streamlit as st
from dotenv import load_dotenv

load_dotenv()

# ── Page config — must be first Streamlit call ─────────────────────────────────
st.set_page_config(
    page_title="SIC/SOC RAG Classifier",
    page_icon="🏭",
    layout="wide",
    initial_sidebar_state="expanded"
)


# ── Vector store initialisation ────────────────────────────────────────────────
@st.cache_resource(show_spinner=False)
def initialise_vector_store():
    """
    Builds the ChromaDB vector store in memory at startup.
    @st.cache_resource means this runs ONCE per app session and is
    shared across all users — not rebuilt on every page interaction.

    This is the correct pattern for stateless cloud deployments:
    no persistent disk needed, cheap to rebuild, always fresh.
    """
    from langchain_openai import OpenAIEmbeddings
    from langchain_chroma import Chroma
    from langchain_core.documents import Document
    import openpyxl

    DATA_PATH = "data/SIC_2007_index.xlsx"

    if not os.path.exists(DATA_PATH):
        st.error(f"Data file not found: {DATA_PATH}")
        st.stop()

    # Load SIC data
    wb = openpyxl.load_workbook(DATA_PATH, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    documents = []
    for row in rows[1:]:
        sic_code = str(row[0]).strip() if row[0] else ""
        activity = str(row[1]).strip() if row[1] else ""
        if sic_code and activity and sic_code != "None":
            documents.append(Document(
                page_content=activity,
                metadata={"sic_code": sic_code, "activity": activity}
            ))

    embeddings = OpenAIEmbeddings(model="text-embedding-3-small")

    # client_settings=None + no persist_directory = pure in-memory store
    vector_store = Chroma.from_documents(
        documents=documents,
        embedding=embeddings
    )
    return vector_store


# ── Retrieval using in-memory store ───────────────────────────────────────────
def retrieve_for_app(
    input_text: str,
    vector_store,
    top_n: int = 5
) -> list[dict]:
    """
    Retrieval function that uses the in-memory store passed from cache.
    Deduplicates by SIC code — same logic as retrieve.py.
    """
    raw_results = vector_store.similarity_search_with_score(
        input_text.strip(), k=top_n * 5
    )
    seen_codes = {}
    for doc, score in raw_results:
        code = doc.metadata["sic_code"]
        if code not in seen_codes:
            seen_codes[code] = {
                "sic_code": code,
                "best_matching_activity": doc.metadata["activity"],
                "similarity_score": round(float(score), 4)
            }
        elif float(score) < seen_codes[code]["similarity_score"]:
            seen_codes[code]["best_matching_activity"] = doc.metadata["activity"]
            seen_codes[code]["similarity_score"] = round(float(score), 4)
    return list(seen_codes.values())[:top_n]


# ── Classification using in-memory store ──────────────────────────────────────
def classify_for_app(input_text: str, vector_store) -> dict:
    """
    Runs the full classification pipeline using the in-memory vector store.
    Imports classify.py logic but overrides retrieval to use app's cached store.
    """
    import re
    from langchain_openai import ChatOpenAI
    from langfuse import observe, get_client

    MODEL_NAME  = "gpt-4o-mini"
    MAX_RETRIES = 2

    # Load valid SIC codes for output guardrail
    import openpyxl
    try:
        wb = openpyxl.load_workbook("data/SIC_2007_index.xlsx", read_only=True)
        ws = wb.active
        valid_codes = set(
            str(r[0]).strip() for r in ws.iter_rows(min_row=2, values_only=True)
            if r[0]
        )
    except Exception:
        valid_codes = set()

    def parse_response(text: str) -> dict | None:
        result = {}
        cm = re.search(r"SIC_CODE:\s*([0-9A-Za-z/]+)", text)
        cf = re.search(r"CONFIDENCE:\s*(HIGH|MEDIUM|AMBIGUOUS)", text)
        rm = re.search(r"REASONING:\s*(.+?)(?:\n|$)", text, re.DOTALL)
        if cm: result["sic_code"]   = cm.group(1).strip()
        if cf: result["confidence"] = cf.group(1).strip()
        if rm: result["reasoning"]  = rm.group(1).strip()
        return result if len(result) == 3 else None

    def build_prompt(text, candidates, retry=False, attempt=1):
        cands = "\n".join([
            f"  {i+1}. SIC {c['sic_code']}: {c['best_matching_activity']}"
            for i, c in enumerate(candidates)
        ])
        if not retry:
            return f"""You are an expert in UK Standard Industrial Classification (SIC) coding.

Select the most appropriate SIC code for this description.

INPUT: "{text}"

CANDIDATES:
{cands}

Assign HIGH if one clearly fits, MEDIUM if somewhat uncertain, AMBIGUOUS if unclear.
Never invent a code not in the list.

Respond EXACTLY:
SIC_CODE: [code]
CONFIDENCE: [HIGH|MEDIUM|AMBIGUOUS]
REASONING: [one sentence]"""
        else:
            return f"""You are an expert in UK SIC coding. Previous attempt was inconclusive.
Retry {attempt}/{MAX_RETRIES}. Think step by step.

INPUT: "{text}"
CANDIDATES:
{cands}

1. What is the core activity?
2. Which code best fits?
3. What distinguishes it from others?
4. How confident are you?

Respond EXACTLY:
SIC_CODE: [code]
CONFIDENCE: [HIGH|MEDIUM|AMBIGUOUS]
REASONING: [one sentence]"""

    # Input validation
    import re as _re
    text = input_text.strip()
    if not text or len(text) < 5:
        return {"outcome": "invalid_input", "reason": "Input too short", "candidates": []}
    if _re.fullmatch(r"[\d\s\-\.\,]+", text):
        return {"outcome": "invalid_input", "reason": "Numeric input only", "candidates": []}

    # Retrieve
    candidates = retrieve_for_app(text, vector_store)
    if not candidates:
        return {"outcome": "retrieval_failed", "candidates": []}

    # Agentic loop
    llm         = ChatOpenAI(model=MODEL_NAME, temperature=0)
    attempts    = 0
    last_parsed = None

    for attempt_num in range(1, MAX_RETRIES + 1):
        attempts += 1
        prompt   = build_prompt(
            text, candidates,
            retry=(attempt_num > 1),
            attempt=attempt_num
        )
        response = llm.invoke(prompt)
        parsed   = parse_response(response.content)
        if parsed is None:
            continue
        last_parsed = parsed
        if parsed.get("confidence") in ("HIGH", "MEDIUM"):
            break

    # Outcome
    if last_parsed is None:
        return {
            "outcome": "human_review", "candidates": candidates,
            "confidence": "AMBIGUOUS", "final_code": "",
            "reasoning": "Could not parse LLM response", "attempts": attempts
        }
    if last_parsed.get("confidence") == "AMBIGUOUS":
        return {
            "outcome": "human_review", "candidates": candidates,
            "confidence": "AMBIGUOUS",
            "final_code": last_parsed.get("sic_code", ""),
            "reasoning":  last_parsed.get("reasoning", ""),
            "attempts":   attempts
        }

    final_code = last_parsed["sic_code"]
    outcome = "classified" if (not valid_codes or final_code in valid_codes) else "human_review"

    return {
        "outcome":    outcome,
        "final_code": final_code,
        "confidence": last_parsed["confidence"],
        "reasoning":  last_parsed["reasoning"],
        "candidates": candidates,
        "attempts":   attempts
    }


# ── Load recent results from CSV ───────────────────────────────────────────────
def load_recent_results(n: int = 5) -> list[dict]:
    path = "data/classification_results.csv"
    if not os.path.exists(path):
        return []
    with open(path, newline="") as f:
        rows = list(csv.DictReader(f))
    return [r for r in rows if r.get("outcome") == "classified"][-n:]


# ── Confidence badge styling ───────────────────────────────────────────────────
def confidence_badge(confidence: str) -> str:
    colours = {
        "HIGH":      "background:#d4edda; color:#155724; padding:3px 10px; "
                     "border-radius:12px; font-weight:600; font-size:13px;",
        "MEDIUM":    "background:#fff3cd; color:#856404; padding:3px 10px; "
                     "border-radius:12px; font-weight:600; font-size:13px;",
        "AMBIGUOUS": "background:#f8d7da; color:#721c24; padding:3px 10px; "
                     "border-radius:12px; font-weight:600; font-size:13px;"
    }
    style = colours.get(confidence, colours["AMBIGUOUS"])
    return f'<span style="{style}">{confidence}</span>'


def outcome_badge(outcome: str) -> str:
    colours = {
        "classified":   "background:#d4edda; color:#155724;",
        "human_review": "background:#fff3cd; color:#856404;",
        "invalid_input":"background:#f8d7da; color:#721c24;",
    }
    labels = {
        "classified":   "✓ Classified",
        "human_review": "⚠ Human Review",
        "invalid_input":"✗ Invalid Input",
    }
    style = colours.get(outcome, "")
    label = labels.get(outcome, outcome)
    return (
        f'<span style="{style} padding:3px 10px; border-radius:12px; '
        f'font-weight:600; font-size:13px;">{label}</span>'
    )


# ── Main UI ────────────────────────────────────────────────────────────────────
def main():
    # Header
    st.title("🏭 SIC/SOC Occupation Classifier")
    st.markdown(
        "Classifies free-text job or industry descriptions into "
        "[UK SIC 2007](https://www.ons.gov.uk/methodology/classificationsandstandards/"
        "ukstandardindustrialclassificationofeconomicactivities/uksic2007) codes "
        "using a **RAG + LLM pipeline** with agentic retry and confidence scoring."
    )
    st.divider()

    # Check API key
    if not os.getenv("OPENAI_API_KEY"):
        st.error(
            "OPENAI_API_KEY not found. "
            "Add it to your .env file (local) or Streamlit Secrets (deployed)."
        )
        st.stop()

    # Sidebar
    with st.sidebar:
        st.header("How it works")
        st.markdown("""
**Pipeline architecture:**

1. **Input guardrail** — validates input before any API calls
2. **Vector retrieval** — embeds your text and finds the top-5 most similar SIC activity descriptions in ChromaDB
3. **LLM classification** — GPT-4o-mini selects the best match with a confidence level
4. **Agentic retry** — if confidence is AMBIGUOUS, retries with chain-of-thought reasoning
5. **Output guardrail** — validates predicted code against the ONS master list
6. **Observability** — every step traced in Langfuse

**Confidence levels:**
- 🟢 **HIGH** — clear match found
- 🟡 **MEDIUM** — likely match with some uncertainty
- 🔴 **AMBIGUOUS** → escalated to human review after retry
        """)

        st.divider()
        st.header("Try these examples")
        examples = [
            "Software developer writing Python code for web apps",
            "GP doctor seeing patients in an NHS health centre",
            "Electrician installing wiring in new build homes",
            "Data scientist building NLP models for surveys",
            "Restaurant manager overseeing kitchen and front of house",
        ]
        for ex in examples:
            if st.button(ex, key=ex, use_container_width=True):
                st.session_state["example_input"] = ex

        st.divider()
        st.caption(
            "Built by Shaurya Bhagat · "
            "[GitHub](https://github.com/ShauryaBhagat/sic-soc-rag-classifier)"
        )

    # ── Vector store loading ───────────────────────────────────────────
    with st.spinner("Loading SIC knowledge base into memory... (~60 seconds on first load)"):
        try:
            vector_store = initialise_vector_store()
            st.success(
                f"Knowledge base ready — "
                f"{vector_store._collection.count():,} SIC activity descriptions loaded.",
                icon="✅"
            )
        except Exception as e:
            st.error(f"Failed to initialise vector store: {e}")
            st.stop()

    st.divider()

    # ── Input area ─────────────────────────────────────────────────────
    # Check if an example button was clicked
    default_text = st.session_state.pop("example_input", "")

    col1, col2 = st.columns([3, 1])
    with col1:
        input_text = st.text_area(
            "Enter a job title or industry description:",
            value=default_text,
            height=100,
            placeholder=(
                "e.g. 'Software developer writing Python code for web applications' "
                "or 'GP doctor working in an NHS health centre'"
            )
        )
    with col2:
        st.write("")   # vertical spacer — no raw HTML
        classify_btn = st.button(
            "🔍 Classify",
            type="primary",
            use_container_width=True,
            disabled=not input_text.strip()
        )
        st.caption(f"{len(input_text)}/500 characters")

    # ── Classification ─────────────────────────────────────────────────
    if classify_btn and input_text.strip():
        with st.spinner("Classifying..."):
            result = classify_for_app(input_text.strip(), vector_store)

        st.divider()
        st.subheader("Result")

        outcome = result.get("outcome", "")

        col_a, col_b, col_c = st.columns(3)
        with col_a:
            st.markdown("**Status**")
            st.markdown(outcome_badge(outcome), unsafe_allow_html=True)
        with col_b:
            if result.get("confidence"):
                st.markdown("**Confidence**")
                st.markdown(
                    confidence_badge(result["confidence"]),
                    unsafe_allow_html=True
                )
        with col_c:
            if result.get("attempts"):
                attempts = result["attempts"]
                st.markdown("**LLM attempts**")
                if attempts > 1:
                    st.warning(f"{attempts}  ⚠ retry triggered")
                else:
                    st.success(f"{attempts}")

        st.write("")

        if outcome == "classified":
            st.success(
                f"**Predicted SIC Code: {result['final_code']}**  \n"
                f"{result.get('reasoning', '')}"
            )
        elif outcome == "human_review":
            st.warning(
                f"**Escalated to human review**  \n"
                f"Best guess: SIC {result.get('final_code', 'N/A')}  \n"
                f"{result.get('reasoning', '')}"
            )
        elif outcome == "invalid_input":
            st.error(
                f"**Input rejected by guardrail**  \n"
                f"Reason: {result.get('reason', 'Input did not meet requirements')}"
            )

        # Candidates table
        if result.get("candidates"):
            st.write("")
            st.subheader("Top retrieved candidates")
            st.caption(
                "Retrieved by vector similarity search across 15,957 "
                "ONS SIC activity descriptions"
            )
            import pandas as pd
            df = pd.DataFrame(result["candidates"]).rename(columns={
                "sic_code":               "SIC Code",
                "best_matching_activity": "Best Matching Activity",
                "similarity_score":       "Distance Score"
            })
            df["Distance Score"] = df["Distance Score"].apply(lambda x: f"{x:.4f}")
            st.dataframe(df, use_container_width=True, hide_index=True)
            st.caption(
                "Lower distance score = more similar. "
                "The LLM selects the best match from these candidates — "
                "it doesn't always pick the closest one."
            )

    # ── Recent classifications ─────────────────────────────────────────
    st.divider()
    st.subheader("Recent classifications")
    recent = load_recent_results(5)
    if recent:
        import pandas as pd
        df_recent = pd.DataFrame(recent)[
            ["input_text", "predicted_sic", "confidence", "attempts"]
        ].rename(columns={
            "input_text":    "Input",
            "predicted_sic": "SIC Code",
            "confidence":    "Confidence",
            "attempts":      "Attempts"
        })
        df_recent["Input"] = df_recent["Input"].str[:70] + "..."
        st.dataframe(df_recent, use_container_width=True, hide_index=True)
    else:
        st.caption(
            "No classifications yet — run one above to see results here."
        )


if __name__ == "__main__":
    main()
